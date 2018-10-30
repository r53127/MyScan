import os
import sqlite3
import win32api

from PyQt5.QtWidgets import QMessageBox
from openpyxl import load_workbook


class StudentDB():
    def __init__(self):
        self.initDB()

    def initDB(self):
        db_path = 'data/student.db'
        # 连接到SQLite数据库
        # 如果文件不存在，会自动在当前目录创建:
        self.conn = sqlite3.connect(db_path)
        # 创建一个Cursor:
        self.cursor = self.conn.cursor()
        # 执行一条SQL语句，创建user表:AUTOINCREMENT类型必须是主键
        self.cursor.execute(
            r'CREATE TABLE IF NOT EXISTS student (userid INTEGER  primary key AUTOINCREMENT ,stuid int,name varchar(50),classid int,classname varchar(20),gender varchar(4))')

    def importStuFromXLS(self, file):
        students = []
        wb = load_workbook(file)
        sheet = wb["Sheet1"]
        # A1格必须是“学号”两个字
        if sheet['A1'].value != '学号':
            QMessageBox.information(None, '提示', '这不是有效的学生库文件！')
            return False
        for row in sheet.rows:  # 使用rows属性，行内第一格从0开始
            student = {}
            # 跳过标题行
            if row[0].value == '学号':
                continue
            # 如果学号不是空
            if row[0].value is not None:
                student['学号'] = row[0].value
                student['姓名'] = row[1].value
                student['班级名称'] = row[2].value
                student['班级代号'] = row[3].value
                students.append(student)
        i=0
        for stud in students:
            # 检查重复
            if self.checkDataByClassID(stud['学号'], stud['班级代号']):
                QMessageBox.information(None, '提示', '该学生重复！' + str(stud['姓名'] + str(stud['学号'])) + str(stud['班级名称']))
                continue
            self.insertDB(stud['学号'], stud['姓名'], stud['班级代号'], stud['班级名称'] )
            i+=1
        return i

    def insertDB(self, stuid, name, classid, classname):
        # 继续执行一条SQL语句，插入一条记录:
        insert_statement = r'insert into student (stuid,name,classid,classname) values (?,?,?,?)'
        self.conn.execute(insert_statement, (stuid, name, classid,classname))
        self.conn.commit()  # 修改类操作必须commit

    def checkDataByClassID(self, stuid, classid):
        query_statement = r"select * from student where stuid='" + str(stuid) + "' and classid='" + str(classid) + "'"
        self.cursor.execute(query_statement)
        return self.cursor.fetchall()

    def checkDataByClassname(self, stuid, classname):
        query_statement = r"select * from student where stuid='" + str(stuid) + "' and classname='" + str(classname) + "'"
        self.cursor.execute(query_statement)
        return self.cursor.fetchall()

    def queryClassname(self):
        query_statement = r"select classname from student"
        self.cursor.execute(query_statement)
        return set(self.cursor.fetchall())

    def queryClassnameByClassID(self, classid):
        query_statement = r"select classname from student where classid='" + str(classid) + "'"
        self.cursor.execute(query_statement)
        return set(self.cursor.fetchall())

    def queryStuByClassname(self, classname):
        query_statement = r"select * from student where classname='" + str(classname) + "' order by stuid ASC"
        self.cursor.execute(query_statement)
        return self.cursor.fetchall()

    def closeDB(self):
        # 关闭Cursor:
        self.cursor.close()


class ScanDB():
    def __init__(self):
        self.initDB()

    def initDB(self):
        db_path = 'data/scan.db'
        # 连接到SQLite数据库
        # 如果文件不存在，会自动在当前目录创建:
        self.conn = sqlite3.connect(db_path)
        # 创建一个Cursor:
        self.cursor = self.conn.cursor()
        # 执行一条SQL语句，创建user表:AUTOINCREMENT类型必须是主键
        self.cursor.execute(
            r'CREATE TABLE IF NOT EXISTS scan (scanid INTEGER  primary key AUTOINCREMENT ,examID varchar(8),classname varchar(20),stuID int,name varchar(50),quesID int,choice varchar(4))')

    def insertDB(self, examid, classname, stuid, name, quesid, choice):
        # 继续执行一条SQL语句，插入一条记录:
        insert_statement = r'insert into scan (examID,classname,stuID,name,quesID,choice) values (?,?,?,?,?,?)'
        self.conn.execute(insert_statement, (examid, classname, stuid, name, quesid, choice))
        self.conn.commit()  # 修改类操作必须commit

    def updateDB(self, stuid, quesid, choice):
        # 继续执行一条SQL语句，插入一条记录:
        update_statement = "update scan set choice='"+choice+"' where quesID='"+str(quesid)+"' and stuID='"+str(stuid)+ "'"
        self.conn.execute(update_statement)
        self.conn.commit()  # 修改类操作必须commit

    def checkData(self, stuid, examid, classname):
        query_statement = r"select * from scan where stuID='" + str(stuid) + "' and examID='" + str(
            examid) + "' and classname='" + str(classname) + "'"
        self.cursor.execute(query_statement)
        return self.cursor.fetchall()

    def queryData(self, classname, examid,quesid,choice):
        query_statement = r"select * from scan where classname='" + str(classname) + "' and examID='" + str(
            examid) + "' and choice like '%" + str(choice) + "%' and quesID='" + str(quesid) + "'"
        self.cursor.execute(query_statement)
        return self.cursor.fetchall()

    def queryCorrectData(self, classname, examid,quesid,ans):
        query_statement = r"select * from scan where classname='" + str(classname) + "' and examID='" + str(
            examid) + "' and choice='" + str(ans) + "' and quesID='" + str(quesid) + "'"
        self.cursor.execute(query_statement)
        return self.cursor.fetchall()

    def queryPersonCount(self, examid, classname):
        query_statement = r"select stuid from scan where examID='" + str(
            examid) + "' and classname='" + str(classname) + "'"
        self.cursor.execute(query_statement)
        return len(set(self.cursor.fetchall()))


    def closeDB(self):
        # 关闭Cursor:
        self.cursor.close()


class ScoreDB():
    def __init__(self):
        self.initDB()

    def initDB(self):
        db_path = 'data/score.db'
        # 连接到SQLite数据库
        # 如果文件不存在，会自动在当前目录创建:
        self.conn = sqlite3.connect(db_path)
        # 创建一个Cursor:
        self.cursor = self.conn.cursor()
        # 执行一条SQL语句，创建user表:AUTOINCREMENT类型必须是主键
        self.cursor.execute(
            r'CREATE TABLE IF NOT EXISTS score (scoreID INTEGER  primary key AUTOINCREMENT ,classname varchar(20),stuID int,name varchar(50),score int,examID varchar(8))')

    def insertDB(self, classname, stuid, name, score, examid):
        # 继续执行一条SQL语句，插入一条记录:
        insert_statement = r'insert into score (classname,stuID,name,score,examID) values (?,?,?,?,?)'
        self.conn.execute(insert_statement, (classname, stuid, name, score, examid))
        self.conn.commit()  # 修改类操作必须commit

    def queryScore(self, classname, examid):
        query_statement = r"select * from score where classname='" + str(classname) + "' and examID='" + str(
            examid) + "'"
        self.cursor.execute(query_statement)
        return self.cursor.fetchall()

    def updateDB(self, stuid,score):
        # 继续执行一条SQL语句，插入一条记录:
        update_statement = "update score set score='"+str(score)+"' where stuID='"+str(stuid)+ "'"
        self.conn.execute(update_statement)
        self.conn.commit()  # 修改类操作必须commit

    def closeDB(self):
        # 关闭Cursor:
        self.cursor.close()


class AnswerDB():
    @classmethod
    def importAnswerFromXLS(cls, file):
        answer = {}
        wb = load_workbook(file)
        sheet = wb["Sheet1"]
        # for i in range(2, sheet.max_row):
        #     answer[sheet["A%d" % i].value] = (sheet["B%d" % i].value, sheet['C%d' % i].value) ## 第一种方法：%字符串占位操作符
        if sheet['A1'].value != '题号':
            QMessageBox.information(None, '提示', '这不是有效的答案文件！')
            return None
        for row in sheet.rows:  # 第二种方法使用rows属性
            if row[0].value == '题号':
                continue
            if row[0].value is not None:
                if row[1].value is not None:
                    answer[row[0].value] = (row[1].value, row[2].value,row[3].value)
                else:
                    answer[row[0].value] = ('', 0,0)
        return answer


class ScoreReportForm():
    def __init__(self):
        self.scoreTemplate = 'data/成绩报表模板.xlsx'
        if not os.path.exists(self.scoreTemplate):
            QMessageBox.information(None, '提示', '找不到报表模板文件！')
            return None
        self.wb = load_workbook(self.scoreTemplate)
        self.sheet = self.wb["成绩表"]
        if self.sheet['A1'].value != '成绩表':
            QMessageBox.information(None, '提示', '这不是有效的模板文件！')

    def makeScoreReport(self, examResults):
        ##examResults内数据分别为scoreID ,classname varchar(20),stuID int,name varchar(20),score int,examID varchar(8))
        examResults = sorted(examResults, key=lambda x: x[1], reverse=True)
        for i, r in enumerate(examResults):
            classname=r[1]
            examid=r[5]
            self.sheet["A%d" % (i + 4)].value =r[2]  # 学号 r[5]  # 时间
            self.sheet["B%d" % (i + 4)].value =r[3]  # 姓名 r[1]  # 班級
            self.sheet["C%d" % (i + 4)].value = r[4]  # 客观分
            if r[0]==0:
                self.sheet["F%d" % (i + 4)].value = '未交卷'  # 如果分数代号为0则显示未交卷

        self.sheet["B2"].value = examid
        self.sheet["E2"].value = classname
        file=r'tmp\\'+classname+examid+r'成绩表.xlsx'
        self.wb.save(file)
        win32api.ShellExecute(0, 'open', file, '', '', 1)


class PaperReportForm():
    def __init__(self):
        self.paperTemplate = 'data/试卷分析模板.xlsx'
        if not os.path.exists(self.paperTemplate):
            QMessageBox.information(None, '提示', '找不到报表模板文件！')
            return None
        self.wb = load_workbook(self.paperTemplate)
        self.sheet=self.wb["试卷分析"]
        if self.sheet['A1'].value != '试卷分析':
            QMessageBox.information(None, '提示', '这不是有效的模板文件！')

    def makePaperReport(self, paperResult):
        ##paperResult内数据：考试时间，班级，题号，正确答案，正确率，A选项率，B选项率，C选项率，D选项率,总人数，总题数
        for i, r in enumerate(paperResult):
            classname=r[1]
            examid=r[0]
            stuCount=r[10]
            quesCount= r[9]
            self.sheet["A%d" % (i + 5)].value = r[2]  # 题号
            self.sheet["B%d" % (i + 5)].value = r[3]  # 正确答案
            self.sheet["C%d" % (i + 5)].value = r[4]  # 正确率
            self.sheet["D%d" % (i + 5)].value = r[5]  # A选项率
            self.sheet["E%d" % (i + 5)].value = r[6]  # B选项率
            self.sheet["F%d" % (i + 5)].value = r[7]  # C选项率
            self.sheet["G%d" % (i + 5)].value = r[8]  # D选项率

        self.sheet["B2"].value = stuCount  # 总人数
        self.sheet["G2"].value = quesCount  # 总题数
        self.sheet["B3"].value = examid  # 考試時間
        self.sheet["G3"].value = classname  # 班級

        file=r'tmp\\'+classname+examid+r'试卷分析.xlsx'
        self.wb.save(file)
        win32api.ShellExecute(0, 'open', file, '', '', 1)

class SaveAsReport():
    def __init__(self):
        self.paperTemplate = 'data/阅卷结果模板.xlsx'
        if not os.path.exists(self.paperTemplate):
            QMessageBox.information(None, '提示', '找不到报表模板文件！')
            return None
        self.wb = load_workbook(self.paperTemplate)
        self.sheet=self.wb["阅卷结果"]
        if self.sheet['A1'].value != '阅卷结果':
            QMessageBox.information(None, '提示', '这不是有效的模板文件！')

    def makeSaveAsReport(self, markingResultView,classname,examid):
        for i, r in enumerate(markingResultView):
            self.sheet["A%d" % (i + 5)].value = r[0]  # 序号
            self.sheet["B%d" % (i + 5)].value =os.path.basename(r[1]) # 文件名
            if r[2]!=0:
                self.sheet["E%d" % (i + 5)].value = r[2][2]  # 分数
                self.sheet["H%d" % (i + 5)].value = str(r[2][1]) # 答题结果
                quesCount= len(r[2][1])
                if r[2][4]==-4:
                    self.sheet["C%d" % (i + 5)].value = r[2][0]  # 学号
                    self.sheet["D%d" % (i + 5)].value = r[2][3]  # 姓名
                    self.sheet["I%d" % (i + 5)].value = '重阅或学号有重复'  # 阅卷结果
                elif r[2][4]==-1:
                    self.sheet["C%d" % (i + 5)].value = None  # 学号
                    self.sheet["D%d" % (i + 5)].value = None  # 姓名
                    self.sheet["I%d" % (i + 5)].value = '未涂学号'  # 阅卷结果
                elif r[2][4]==-2:
                    self.sheet["C%d" % (i + 5)].value = r[2][0]  # 学号
                    self.sheet["D%d" % (i + 5)].value = None  # 姓名
                    self.sheet["I%d" % (i + 5)].value = '班级冲突'  # 阅卷结果
                elif r[2][4]==-3:
                    self.sheet["C%d" % (i + 5)].value = r[2][0]  # 学号
                    self.sheet["D%d" % (i + 5)].value = None  # 姓名
                    self.sheet["I%d" % (i + 5)].value = '学号不存在'  # 阅卷结果
                elif r[2][4] == 1:
                    self.sheet["C%d" % (i + 5)].value = r[2][0]  # 学号
                    self.sheet["D%d" % (i + 5)].value = r[2][3]  # 姓名
                    self.sheet["I%d" % (i + 5)].value = '正常'  # 阅卷结果
            else:
                self.sheet["I%d" % (i + 5)].value = '答题卡无法识别'  # 阅卷结果

        self.sheet["B2"].value =  quesCount##总题数
        self.sheet["I2"].value = len(markingResultView)  # # 总人数
        self.sheet["B3"].value = examid  # 考試時間
        self.sheet["I3"].value = classname  # 班級

        file=r'tmp\\'+classname+examid+r'阅卷结果.xlsx'
        self.wb.save(file)
        win32api.ShellExecute(0, 'open', file, '', '', 1)

if __name__ == "__main__":
    test = ScoreDB()
    test.updateDB(26,100)
